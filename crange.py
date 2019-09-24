#ライブラリ
import cv2 #映像処理
import numpy as np
from scipy.interpolate import interp1d # scipyのモジュール
import math
import openpyxl #excelファイル読み込み
import matplotlib.pyplot as plt #グラフ作成
#from scipy.optimize import curve_fit 
from scipy import optimize #近似

show = bool(True) #画面表示の有無 True: 画面表示あり
outputpng = bool(True) #画像アウトプット True:あり
outputexcel = bool(False) #Excelへのアウトプット True:あり

#カラートラッキング
def color_detect(img):
    # HSV色空間に変換
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    # 色のHSVの値域
    hsv_min = np.array([160,100,50])
    hsv_max = np.array([180,255,255])
    mask = cv2.inRange(hsv, hsv_min, hsv_max)
    return mask

print('OpenCV'+cv2.__version__)

VIDEOFILE = get_list(sheet, 1, 2, maxcol+1) #ビデオファイル名


#メイン
def main():
    #まとめデータ作成用
    allgraphdata_pos = [[],[]]
    allgraphtheory_pos = [[],[]]
    allgraphdata_vel = [[],[]]
    allgraphtheory_vel = [[],[]]
    allgraphdata_acc = [[],[]]
    allgraphtheory_acc = [[],[]]
    allgraphdata_Fv = [[],[]]
    allgraphtheory_Fv = [[],[]]
    Fvall = [[],[]]
    Fvall_fit = []
    pos_model = [[],[]]
    vel_model = [[],[]]
    acc_model = [[],[]]

    tmp = []

    for h in range(len(VIDEOFILE)): #作業リストにかかれている全ての動画の解析をする
        if VIDEOFILE[h] != None:
            #ビデオ
            video = cv2.VideoCapture(VIDEOFILE[h]+'.mp4')
            
            data = []
            t = 0
            allframe = int(video.get(cv2.CAP_PROP_FRAME_COUNT)) #総フレーム数

            #動画のカラートラッキング等の処理
            for i in range(allframe):
                # フレームを取得
                ret, frame = video.read()

                # 色検出
                mask = color_detect(frame)

                #最大面積の重心
                x, y = calc_max_point(mask)

                #単位変換
                x *= adjlen[h]
                y *= adjlen[h]

                #屈折処理
                x = refraction(x,h)

                #データの追加
                data.append([t, x, y])

                # 結果表示
                if show == True:
                    cv2.circle(frame, (int(x / adjlen[h]), int(y / adjlen[h])), 20, (0, 0, 255), 10)
                    cv2.imshow("Frame", frame)
                    cv2.imshow("Mask", mask)

                t = video.get(cv2.CAP_PROP_POS_FRAMES) / FPS[h]

                # qキーが押されたら途中終了
                if cv2.waitKey(25) & 0xFF == ord('q'):
                    break
                
                #np.savetxt("data.csv", np.array(data), delimiter=",")
            cv2.destroyAllWindows()

            #解析結果の処理

            #位置=生データ   x成分のみ使用
            dat = [[x[0] for x in data], [x[1] for x in data]]

            #外れ値処理
            dat[0], dat[1] = outlier_detection(dat[1], dat[0], 0.5, 0.001, 45)[0], outlier_detection(dat[1], dat[0], 0.5, 0.001, 45)[1]
            
            #平滑化
            dat[0], dat[1] = convolve(dat[0],10), convolve(dat[1],10)

            #位置の配列の最初を0にそろえる
            for i in reversed(range(len(dat[1]))):
                dat[1][i] -= dat[1][0]

            #速度 
            vel = [[],[]]
            vel[0], vel[1] = differential_calculus(dat[1], dat[0])[0], differential_calculus(dat[1], dat[0])[1]

            #平滑化
            vel[0], vel[1] = convolve(vel[0],10), convolve(vel[1],10)
            
            #加速度
            acc = [[],[]]
            acc[0], acc[1] = differential_calculus(vel[1], vel[0])[0], differential_calculus(vel[1], vel[0])[1]

            #平滑化
            acc[0], acc[1] = convolve(acc[0],10), convolve(acc[1],10)

            #v-aグラフ作成用処理
            av = [[],[]]
            t = 0
            for i in range(len(acc[0])):
                #jを求める(tに代入)
                for j in range(t, len(vel[0])):
                    if vel[0][j] > acc[0][i]:
                        t = j
                        break
                #procacc = acc[1][i-1] + (vel[0][t] - acc[0][i-1]) / (acc[0][i] - acc[0][i-1]) * (acc[1][i] - acc[1][i-1])
                procacc = ((vel[0][t]-acc[0][i-1])*acc[1][i] + (acc[0][i]-vel[0][t])*acc[1][i-1]) / (acc[0][i]-acc[0][i-1])
                av[0].append(vel[1][t])
                av[1].append(procacc)

            #F-vグラフ作成用処理
            Fv = [[],[]]
            for i in range(len(av[0])):
                Fv[0].append(av[0][i])
                Fv[1].append(allvolume[h] * (rhoO[h]-rhoW[h]) * g - allmass[h] *av[1][i] - 1 / 4 * math.pi * rhoW[h] * ballradius[h] ** 2 * av[0][i] ** 2)


            

            #サンプル数の表示
            '''
            print(VIDEOFILE[h])
            print('datlen','\t',len(dat[1]),sep='')
            print('vellen','\t',len(vel[1]),sep='')
            print('acclen','\t',len(acc[1]),sep='')
            '''

            #まとめグラフ作成用データ
            allgraphdata_pos[0].append(dat[0])
            allgraphdata_pos[1].append(dat[1])
            allgraphdata_vel[0].append(vel[0])
            allgraphdata_vel[1].append(vel[1])
            allgraphdata_acc[0].append(acc[0])
            allgraphdata_acc[1].append(acc[1])
            allgraphdata_Fv[0].append(Fv[0])
            allgraphdata_Fv[1].append(Fv[1])



            #excelに保存
            
            colnum = 10
            sheet = wb['解析結果']
            sheet.cell(row=1, column=h*colnum+1, value='ビデオ名称')
            sheet.cell(row=1, column=h*colnum+2, value='位置-時間')
            sheet.cell(row=1, column=h*colnum+3, value='位置')
            sheet.cell(row=1, column=h*colnum+4, value='速度-時間')
            sheet.cell(row=1, column=h*colnum+5, value='速度')
            sheet.cell(row=1, column=h*colnum+6, value='加速度-時間')
            sheet.cell(row=1, column=h*colnum+7, value='加速度')
            sheet.cell(row=1, column=h*colnum+8, value='速度-with加速度')
            sheet.cell(row=1, column=h*colnum+9, value='加速度-with速度')
            sheet.cell(row=2, column=h*colnum+1, value=VIDEOFILE[h])
            write_list(dat[0], sheet, 2, h*colnum+2)
            write_list(dat[1], sheet, 2, h*colnum+3)
            write_list(vel[0], sheet, 2, h*colnum+4)
            write_list(vel[1], sheet, 2, h*colnum+5)
            write_list(acc[0], sheet, 2, h*colnum+6)
            write_list(acc[1], sheet, 2, h*colnum+7)
            write_list(Fv[0], sheet, 2, h*colnum+8)
            write_list(Fv[1], sheet, 2, h*colnum+9)
            wb.save('物理チャレンジ2019.xlsx')
            







        if VIDEOFILE[h] == None:
            #全グラフのそれぞれの平均値をとった代表グラフの作成
            timlen = 0.005#この秒数の間を平均する
            pos_model = model_make(timlen,allgraphdata_pos)
            vel_model = model_make(timlen,allgraphdata_vel)
            acc_model = model_make(timlen,allgraphdata_acc)
            Fv_model = model_make(0.05,allgraphdata_Fv)

            #F-vの2次式での近似
            #print(allgraphdata_Fv[0])
            fv = [[],[]]
            for i in range(len(allgraphdata_Fv[0])):
                for j in range(len(allgraphdata_Fv[0][i])):
                    fv[0].append(allgraphdata_Fv[0][i][j])
                    fv[1].append(allgraphdata_Fv[1][i][j])
            param, cov = optimize.curve_fit(linear_fit, fv[0], fv[1])
            print(param[0])
            Fvall_fit = [[],[]]
            for i in range(len(allgraphdata_Fv[0][0])):
                Fvall_fit[0].append(allgraphdata_Fv[0][0][i])
                Fvall_fit[1].append(param * allgraphdata_Fv[0][0][i] ** 2)
            
            theta = 30 * math.pi / 180
            #mu = 1/2 * 0.642 * math.pi * theta * math.sqrt(cylinderlengh[h-1] / objradius[h-1]) * rhoW[h-1] * objradius[h-1] ** 2 / (allmass[h-1]-ballmass[h-1])
            mu = 1/2 * math.pi * math.sqrt(theta) * math.sqrt(objradius[h-1] / cylinderlengh[h-1]) * rhoW[h-1] * objradius[h-1] ** 2 / allmass[h-1]
            # + 1/4 * math.pi * rhoW[h-1] * ballradius[h-1] ** 2 / ballmass[h-1]
            #MU = 1/2 * 0.642 * math.pi * theta * math.sqrt(cylinderlengh[h-1] / objradius[h-1]) * rhoW[h-1] * objradius[h-1] ** 2
            MU = 1/2 * math.pi * math.sqrt(theta) * math.sqrt(objradius[h-1] / cylinderlengh[h-1]) * rhoW[h-1] * objradius[h-1] ** 2

            #慣性抵抗と粘性抵抗を考慮した運動方程式の実装(F-vグラフ)
            visineResistance = [[],[]]
            for i in range(len(allgraphdata_Fv[0][0])):
                visineResistance[0].append(allgraphdata_Fv[0][0][i])
                visineResistance[1].append(MU * (allgraphdata_Fv[0][0][i]) ** 2)

            #慣性抵抗を考慮したa-tグラフ
            aine = [[],[]]
            for i in range(len(allgraphdata_acc[0][0])):
                aine[0].append(allgraphdata_acc[0][0][i])
                aine[1].append((rhoO[h-1] - rhoW[h-1]) / rhoO[h-1] * g * (1-(math.tanh(math.sqrt(mu * (rhoO[h-1] - rhoW[h-1]) / rhoO[h-1] * g) * allgraphdata_acc[0][0][i])) ** 2))

            #慣性抵抗を考慮したv-tグラフ
            vine = [[],[]]
            for i in range(len(allgraphdata_vel[0][0])):
                vine[0].append(allgraphdata_vel[0][0][i])
                vine[1].append(math.sqrt((rhoO[h-1] - rhoW[h-1]) / mu / rhoO[h-1] * g) * math.tanh(math.sqrt(mu * (rhoO[h-1] - rhoW[h-1]) / rhoO[h-1] * g) * allgraphdata_vel[0][0][i]))

            #慣性抵抗を考慮したx-tグラフ
            xine = [[],[]]
            for i in range(len(allgraphdata_pos[0][0])):
                xine[0].append(allgraphdata_pos[0][0][i])
                xine[1].append(1 / mu * math.log(math.cosh(math.sqrt(mu * (rhoO[h-1] - rhoW[h-1]) / rhoO[h-1] * g) * allgraphdata_pos[0][0][i])))
            
            #全ての実験をまとめたグラフ
            fig = plt.figure(figsize=(20,12))
            plt.rcParams["font.family"] = "IPAexGothic" # フォントの種類
            fig.text(0.05,0.95,objname[h-1]+' 実験日:'+str(date[h-1].date())+' 水温:'+str(temperature[h-1])+'℃',fontsize=25)
            plt.subplots_adjust(left=0.05, right=0.98, bottom=0.05, top=0.9, wspace=0.15, hspace=0.2)
            linwid = 2
            
            #左上
            plt.subplot(2,2,1)
            for i in range(0,len(allgraphdata_pos[0])):
                plt.plot(allgraphdata_pos[0][i], allgraphdata_pos[1][i], '.', label=VIDEOFILE[h-3+i])
            plt.plot(allgraphtheory_pos[0], allgraphtheory_pos[1], '-', label='位置-理論値',lw=linwid)
            plt.plot(pos_model[0], pos_model[1], '-', label="位置-平均値",lw=linwid)
            plt.plot(xine[0], xine[1], '-', label="位置-理論値",lw=linwid)
            plt.xlim(0,0.35) #x軸の最小値、最大値
            plt.ylim(0,)#0.125) #y軸の最小値を0に設定
            plt.title("時間と位置の関係", fontsize=20) #題名
            plt.xlabel("時間[秒]", fontsize=18)     #x軸ラベル
            plt.ylabel("位置[m]", fontsize=18)    #y軸ラベル
            plt.grid() #グリッド表示
            plt.legend(fontsize=18) #凡例表示
            plt.tick_params(labelsize=15)

            # 右上
            plt.subplot(2,2,2)
            for i in range(0,len(allgraphdata_vel[0])):
                plt.plot(allgraphdata_vel[0][i], allgraphdata_vel[1][i], '.', label=VIDEOFILE[h-3+i])
            plt.plot(allgraphtheory_vel[0], allgraphtheory_vel[1], '-', label='速度-理論値',lw=linwid)
            plt.plot(vel_model[0], vel_model[1], '-', label="速度-平均値",lw=linwid)
            plt.plot(vine[0], vine[1], '-', label="速度-理論値",lw=linwid)
            plt.xlim(0,0.35) #x軸の最小値、最大値
            plt.ylim(0,1) #y軸の最小値を0に設定
            plt.title("時間と速度の関係", fontsize=20) #題名
            plt.xlabel("時間[秒]", fontsize=18)     #x軸ラベル
            plt.ylabel("速度[m/s]", fontsize=18)    #y軸ラベル
            plt.grid() #グリッド表示
            plt.legend(fontsize=18) #凡例表示
            plt.tick_params(labelsize=15)

            # 左下
            plt.subplot(2,2,3)
            for i in range(0,len(allgraphdata_Fv[0])):
                plt.plot(allgraphdata_acc[0][i], allgraphdata_acc[1][i], '.', label=VIDEOFILE[h-3+i])
            plt.plot(allgraphtheory_acc[0], allgraphtheory_acc[1], '-', label='加速度-理論値',lw=linwid)
            plt.plot(acc_model[0], acc_model[1], '-', label="加速度-平均値",lw=linwid)
            plt.plot(aine[0], aine[1], '-', label="加速度-理論値",lw=linwid)
            plt.xlim(0,)#0.35) #x軸の最小値、最大値
            plt.ylim(-5,20) #y軸の最小値
            plt.title("時間と加速度の関係", fontsize=20) #題名
            plt.xlabel("時間[秒]", fontsize=18)     #x軸ラベル
            plt.ylabel("加速度[m/s^2]", fontsize=18)    #y軸ラベル
            plt.grid() #グリッド表示
            plt.legend(fontsize=18,ncol=3) #凡例表示
            plt.tick_params(labelsize=15)

            # 右下
            plt.subplot(2,2,4)
            for i in range(0,len(allgraphdata_Fv[0])):
                plt.plot(allgraphdata_Fv[0][i], allgraphdata_Fv[1][i], '.', label=VIDEOFILE[h-3+i])
            plt.plot(allgraphtheory_Fv[0], allgraphtheory_Fv[1], '-', label='抵抗力-理論値',lw=linwid)
            plt.plot(Fv_model[0], Fv_model[1], '-', label="抵抗力-平均値",lw=linwid)
            plt.plot(visineResistance[0], visineResistance[1], '-', label="抵抗力-理論値",lw=linwid)
            plt.plot(Fvall_fit[0], Fvall_fit[1], '-', label="近似値" + str('{:.3f}'.format(param[0])),lw=linwid)
            plt.xlim(0,1) #x軸の最小値を0に設定
            plt.ylim(-0.04,0.06) #y軸の最小値
            plt.title("速度と抵抗力の関係", fontsize=20) #題名
            plt.xlabel("速度[m/s]", fontsize=18)     #x軸ラベル
            plt.ylabel("抵抗力[N]", fontsize=18)    #y軸ラベル
            plt.grid() #グリッド表示
            plt.legend(fontsize=16,ncol=3) #凡例表示
            plt.tick_params(labelsize=15)

            if outputpng == True:
                plt.savefig(objname[h-1]+'.png')
            if show == True:
                plt.show()
            
            allgraphdata_pos = [[],[]]
            allgraphtheory_pos = [[],[]]
            allgraphdata_vel = [[],[]]
            allgraphtheory_vel = [[],[]]
            allgraphdata_acc = [[],[]]
            allgraphtheory_acc = [[],[]]
            allgraphdata_Fv = [[],[]]
            allgraphtheory_Fv = [[],[]]





        #excelに保存
        colnum = 10
        if outputexcel == True:
            if(h == len(VIDEOFILE)-1):
                sheet = wb['近似出力']
                sheet.cell(row=1, column=1, value='速度')
                sheet.cell(row=1, column=2, value='抵抗力')
                write_list(Fv_model[0], sheet, 2, 1)
                write_list(Fv_model[1], sheet, 2, 2)

                sheet.cell(row=1, column=3, value='時間')
                sheet.cell(row=1, column=4, value='位置')
                sheet.cell(row=1, column=5, value='時間')
                sheet.cell(row=1, column=6, value='速度')
                sheet.cell(row=1, column=7, value='時間')
                sheet.cell(row=1, column=8, value='加速度')
                write_list(pos_model[0], sheet, 2, 3)
                write_list(pos_model[1], sheet, 2, 4)
                write_list(vel_model[0], sheet, 2, 5)
                write_list(vel_model[1], sheet, 2, 6)
                write_list(acc_model[0], sheet, 2, 7)
                write_list(acc_model[1], sheet, 2, 8)

        if outputexcel == True:
            wb.save('物理チャレンジ2019.xlsx')


if __name__ == '__main__':
    main()